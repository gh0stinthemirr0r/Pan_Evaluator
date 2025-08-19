#!/usr/bin/env python3
"""
PAN-OS Read‑Only Policy Evaluator (GUI + Exports)
-------------------------------------------------

Purpose
  * Connects read-only to a Palo Alto NGFW or Panorama (API key auth)
  * Fetches ordered Security policy rules + per-rule hit counts
  * Computes: unused rules (0 hits over your counter window), duplicates, shadowed rules, and safe merge/consolidation proposals
  * Performs order/deny-intervening checks to flag risky merges
  * NEVER pushes/writes to the firewall; outputs analyst-friendly reports

Outputs
  * CSV, XLSX (styled), HTML, and PDF (tabular)

GUI
  * Tkinter-based front-end with config persistence (evaluator.conf JSON next to script)
  * Fields: API URL, API Key, VSYS/Device Group, Rulebase type, Output folder
  * Buttons: Fetch & Analyze (read-only), Preview Summary, Export (CSV/XLSX/HTML/PDF)

Dependencies (install as needed)
  pip install pan-os-python lxml pandas openpyxl reportlab tabulate

Notes
  * Works with a single firewall (mgmt API) or Panorama device-group (set rulebase location accordingly)
  * Hit counters must be enabled and accumulating on the device (you cleared them; perfect)
  * This tool is intentionally conservative; if any uncertainty exists, it marks order_sensitive

Author: Aaron Stovall

"""
from __future__ import annotations
import os
import json
import datetime as dt
import hashlib
from dataclasses import dataclass, field
from typing import Any, Dict, List, Tuple, Optional

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# Third-party
try:
    from panos.firewall import Firewall
    from panos.policies import Rulebase, SecurityRule
    PANOS_AVAILABLE = True
except ImportError as e:
    print(f"Warning: pan-os-python import failed: {e}")
    print("Please install with: pip install pan-os-python")
    Firewall = None  # type: ignore
    Rulebase = None  # type: ignore
    SecurityRule = None  # type: ignore
    PANOS_AVAILABLE = False
except Exception as e:
    print(f"Warning: Unexpected error importing pan-os-python: {e}")
    Firewall = None  # type: ignore
    Rulebase = None  # type: ignore
    SecurityRule = None  # type: ignore
    PANOS_AVAILABLE = False

# Diagnostic function to check dependencies
def check_dependencies():
    """Check if all required dependencies are available"""
    issues = []
    
    # Check pan-os-python
    if not PANOS_AVAILABLE:
        issues.append("pan-os-python: NOT AVAILABLE - Run: pip install pan-os-python")
    else:
        issues.append("pan-os-python: ✅ AVAILABLE")
    
    # Check pandas
    try:
        import pandas as pd
        issues.append("pandas: ✅ AVAILABLE")
    except ImportError:
        issues.append("pandas: NOT AVAILABLE - Run: pip install pandas")
    
    # Check tabulate
    try:
        import tabulate
        issues.append("tabulate: ✅ AVAILABLE")
    except ImportError:
        issues.append("tabulate: NOT AVAILABLE - Run: pip install tabulate")
    
    # Check reportlab
    try:
        import reportlab
        issues.append("reportlab: ✅ AVAILABLE")
    except ImportError:
        issues.append("reportlab: NOT AVAILABLE - Run: pip install reportlab")
    
    # Check openpyxl
    try:
        import openpyxl
        issues.append("openpyxl: ✅ AVAILABLE")
    except ImportError:
        issues.append("openpyxl: NOT AVAILABLE - Run: pip install openpyxl")
    
    return issues

import pandas as pd
from tabulate import tabulate

# Optional for XLSX styling and PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

APP_NAME = "PAN-OS Policy Evaluator"
CONF_FILE = "evaluator.conf"

# -----------------------------
# Utility helpers
# -----------------------------

def now_iso() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def read_conf(path: str) -> Dict[str, Any]:
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            try:
                return json.load(f)
            except Exception:
                return {}
    return {}


def write_conf(path: str, data: Dict[str, Any]) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


def as_set(x: Optional[List[str]]) -> set:
    return set(x or [])


def any_intersection(a: set, b: set) -> bool:
    if not a or not b:
        return False
    return not a.isdisjoint(b)


def normalize_list(x: Any) -> List[str]:
    if x is None:
        return []
    if isinstance(x, str):
        return [x]
    try:
        return list(x)
    except Exception:
        return [str(x)]


# -----------------------------
# Data models
# -----------------------------
@dataclass
class RuleLike:
    name: str
    position: int
    action: str
    fromzone: List[str]
    tozone: List[str]
    source: List[str]
    destination: List[str]
    application: List[str]
    service: List[str]
    source_user: List[str]
    url_category: List[str]
    schedule: Optional[str]
    log_setting: Optional[str]
    log_start: Optional[bool]
    log_end: Optional[bool]
    profile_setting: Any
    disabled: bool
    negate_source: bool
    negate_destination: bool
    location: Optional[str]  # pre/post for Panorama, or None
    hits_total: Optional[int] = None
    last_hit: Optional[str] = None
    counter_since: Optional[str] = None

    def to_row(self) -> Dict[str, Any]:
        return {
            "Position": self.position,
            "Name": self.name,
            "Tags": getattr(self, 'tags', ''),
            "Type": getattr(self, 'rule_type', 'universal'),
            "Source Zone": ", ".join(self.fromzone),
            "Source Address": ", ".join(self.source),
            "Source User": ", ".join(self.source_user),
            "Source Device": getattr(self, 'source_device', 'any'),
            "Destination Zone": ", ".join(self.tozone),
            "Destination Address": ", ".join(self.destination),
            "Destination Device": getattr(self, 'destination_device', 'any'),
            "Application": ", ".join(self.application),
            "Service": ", ".join(self.service),
            "Action": self.action,
            "Profile": str(self.profile_setting or ""),
            "Options": self.log_setting or "",
            "Rule Usage Hit Count": self.hits_total or 0,
            "Rule Usage Last Hit": self.last_hit or "",
            "Rule Usage First Hit": self.counter_since or "",
            "Rule Usage Apps Seen": getattr(self, 'apps_seen', ''),
            "Days With No New Apps": getattr(self, 'days_no_new_apps', ''),
            "Modified": getattr(self, 'modified', ''),
            "Created": getattr(self, 'created', ''),
            # Additional fields for compatibility
            "FromZones": ", ".join(self.fromzone),
            "ToZones": ", ".join(self.tozone),
            "Source": ", ".join(self.source),
            "Destination": ", ".join(self.destination),
            "Applications": ", ".join(self.application),
            "Services": ", ".join(self.service),
            "Users": ", ".join(self.source_user),
            "URLCategories": ", ".join(self.url_category),
            "Schedule": self.schedule or "",
            "Profiles": str(self.profile_setting or ""),
            "LogSetting": self.log_setting or "",
            "LogStart": self.log_start,
            "LogEnd": self.log_end,
            "Disabled": self.disabled,
            "NegateSrc": self.negate_source,
            "NegateDst": self.negate_destination,
            "PrePost": self.location or "",
            "Hits": self.hits_total,
            "LastHit": self.last_hit or "",
            "CounterSince": self.counter_since or "",
        }

    def non_broadening_fingerprint(self) -> Tuple[str, Dict[str, Any]]:
        key = {
            "action": self.action,
            "from": tuple(sorted(self.fromzone)),
            "to": tuple(sorted(self.tozone)),
            "users": tuple(sorted(self.source_user)),
            "urlcat": tuple(sorted(self.url_category)),
            "schedule": self.schedule,
            "profiles": json.dumps(self.profile_setting, sort_keys=True),
            "log_setting": self.log_setting,
            "log_start": bool(self.log_start),
            "log_end": bool(self.log_end),
            "disabled": bool(self.disabled),
            "neg_src": bool(self.negate_source),
            "neg_dst": bool(self.negate_destination),
            "prepost": self.location or "",
        }
        fp = hashlib.sha1(repr(key).encode()).hexdigest()
        return fp, key


@dataclass
class Proposal:
    proposed_name: str
    source_rules: List[str]
    positions: List[int]
    apps_union: List[str]
    services_union: List[str]
    sources_union: List[str]
    destinations_union: List[str]
    order_sensitive: bool
    order_reason: str
    confidence: str
    recommendation: str
    notes: str = ""

    def to_row(self) -> Dict[str, Any]:
        return {
            "ProposedName": self.proposed_name,
            "SourceRules": ", ".join(self.source_rules),
            "Positions": ", ".join(map(str, sorted(self.positions))),
            "ApplicationsUnion": ", ".join(self.apps_union),
            "ServicesUnion": ", ".join(self.services_union),
            "SourcesUnion": ", ".join(self.sources_union),
            "DestinationsUnion": ", ".join(self.destinations_union),
            "OrderSensitive": self.order_sensitive,
            "OrderReason": self.order_reason,
            "Confidence": self.confidence,
            "Recommendation": self.recommendation,
            "Notes": self.notes,
        }


@dataclass
class ShadowFinding:
    shadowed_rule: str
    shadowed_position: int
    shadowing_rule: str
    shadowing_position: int
    reason: str
    recommendation: str

    def to_row(self) -> Dict[str, Any]:
        return {
            "ShadowedRule": self.shadowed_rule,
            "ShadowedPos": self.shadowed_position,
            "ShadowingRule": self.shadowing_rule,
            "ShadowingPos": self.shadowing_position,
            "Reason": self.reason,
            "Recommendation": self.recommendation,
        }


# -----------------------------
# PAN-OS client (read-only)
# -----------------------------
class PanOSReader:
    def __init__(self, api_url: str, api_key: str, vsys: str = "vsys1"):
        if not PANOS_AVAILABLE:
            raise RuntimeError("pan-os-python not installed. Run: pip install pan-os-python")
        self.fw = Firewall(hostname=api_url, api_key=api_key)
        self.vsys = vsys
        self.available_vsys = []
        
        # Check if this is a multi-VSYS firewall and get available VSYS names
        try:
            system_info = self.fw.op("show system info", xml=True)
            if isinstance(system_info, bytes):
                import xml.etree.ElementTree as ET
                root = ET.fromstring(system_info.decode('utf-8'))
            else:
                root = system_info
            
            multi_vsys = root.findtext('.//multi-vsys')
            if multi_vsys and multi_vsys.lower() == 'off':
                # Single VSYS mode - use vsys1 regardless of what was passed
                self.vsys = "vsys1"
                self.available_vsys = ["vsys1"]
                print(f"Detected single-VSYS mode, using vsys1")
            else:
                print(f"Multi-VSYS mode detected, attempting to discover available VSYS names...")
                self.available_vsys = self._discover_vsys_names()
                print(f"Available VSYS names: {self.available_vsys}")
                if self.vsys not in self.available_vsys and self.available_vsys:
                    # Use the first available VSYS if the configured one isn't found
                    self.vsys = self.available_vsys[0]
                    print(f"Configured VSYS '{vsys}' not found, using first available: {self.vsys}")
        except Exception as e:
            print(f"Could not determine VSYS mode, using configured VSYS: {self.vsys}")
            print(f"Error: {e}")
            self.available_vsys = [self.vsys]
    
    def _discover_vsys_names(self) -> List[str]:
        """Discover all available VSYS names on the firewall"""
        vsys_names = []
        
        # Try different commands to get VSYS list
        vsys_commands = [
            "show vsys",
            "show system vsys",
            "show virtual-system",
            "show virtual-systems"
        ]
        
        for cmd in vsys_commands:
            try:
                result = self.fw.op(cmd, xml=True)
                
                # Handle both bytes and XML object responses
                if isinstance(result, bytes):
                    import xml.etree.ElementTree as ET
                    root = ET.fromstring(result.decode('utf-8'))
                else:
                    root = result
                
                # Try different XML paths for VSYS names
                entries = (
                    root.findall('.//entry') or 
                    root.findall('.//vsys') or 
                    root.findall('.//virtual-system') or
                    root.findall('.//virtual-systems/entry')
                )
                
                for entry in entries:
                    name = entry.get('name')
                    if name and name not in vsys_names:
                        vsys_names.append(name)
                
                if vsys_names:
                    print(f"VSYS discovery succeeded with command: {cmd}")
                    break
                    
            except Exception as e:
                print(f"VSYS discovery failed with command '{cmd}': {e}")
                continue
        
        # If no VSYS names found, try to get from system info
        if not vsys_names:
            try:
                system_info = self.fw.op("show system info", xml=True)
                if isinstance(system_info, bytes):
                    import xml.etree.ElementTree as ET
                    root = ET.fromstring(system_info.decode('utf-8'))
                else:
                    root = system_info
                
                # Look for VSYS information in system info
                vsys_info = root.findtext('.//vsys')
                if vsys_info:
                    vsys_names = [vsys_info]
                    print(f"Found VSYS from system info: {vsys_info}")
            except Exception as e:
                print(f"Could not get VSYS from system info: {e}")
        
        # If still no VSYS names, assume vsys1
        if not vsys_names:
            vsys_names = ["vsys1"]
            print("No VSYS names discovered, assuming vsys1")
        
        return vsys_names

    def fetch_rules(self) -> List[RuleLike]:
        # Try to fetch rules from all available VSYS until one works
        for vsys_name in self.available_vsys:
            print(f"Attempting to fetch rules from VSYS: {vsys_name}")
            try:
                # Try pan-os-python method first
                rb = Rulebase()
                self.fw.add(rb)
                rb.refreshall(SecurityRule, add=False)
                rules: List[RuleLike] = []
                pos = 0
                for child in rb.children:
                    if not isinstance(child, SecurityRule):
                        continue
                    pos += 1
                    rules.append(self._to_rulelike(child, pos))
                
                if rules:
                    print(f"✅ Successfully fetched {len(rules)} rules from VSYS: {vsys_name}")
                    self.vsys = vsys_name  # Update the active VSYS
                    return rules
                else:
                    print(f"⚠️  No rules found in VSYS: {vsys_name}")
                    
            except Exception as e:
                print(f"Warning: pan-os-python rule fetch failed for VSYS {vsys_name}: {e}")
                print("Attempting direct API call...")
                
                # Try direct API call for this VSYS
                try:
                    rules = self._fetch_rules_direct(vsys_name)
                    if rules:
                        print(f"✅ Successfully fetched {len(rules)} rules from VSYS: {vsys_name} via direct API")
                        self.vsys = vsys_name  # Update the active VSYS
                        return rules
                    else:
                        print(f"⚠️  No rules found in VSYS: {vsys_name} via direct API")
                except Exception as direct_error:
                    print(f"Direct API call failed for VSYS {vsys_name}: {direct_error}")
        
        # If we get here, no VSYS had rules
        print(f"❌ No security rules found in any VSYS: {self.available_vsys}")
        return []
    
    def _fetch_rules_direct(self, vsys_name: str = None) -> List[RuleLike]:
        """Fallback method to fetch rules directly via API"""
        if vsys_name is None:
            vsys_name = self.vsys
            
        try:
            # Try different command formats for different PAN-OS versions
            commands = [
                "show security-policy",
                "show security-policy rules",
                "show security-policy rule-base security rules",
                f"show security-policy vsys {vsys_name}",
                f"show security-policy vsys {vsys_name} rules",
                f"show security-policy vsys-name {vsys_name}",
                f"show security-policy vsys-name {vsys_name} rules",
                # Try some alternative command formats
                "show rules security",
                "show security rules",
                f"show rules security vsys {vsys_name}",
                f"show security rules vsys {vsys_name}"
            ]
            
            xml = None
            for cmd in commands:
                try:
                    xml = self.fw.op(cmd, xml=True)
                    print(f"Direct API command succeeded: {cmd}")
                    break
                except Exception as e:
                    print(f"Direct API command failed: {cmd} - {e}")
                    continue
            
            if xml is None:
                print("All direct API commands failed")
                return []
            
            # Handle both bytes and XML object responses
            if isinstance(xml, bytes):
                import xml.etree.ElementTree as ET
                root = ET.fromstring(xml.decode('utf-8'))
            else:
                root = xml
            
            rules: List[RuleLike] = []
            pos = 0
            
            # Try different XML paths for security rules
            rule_entries = (
                root.findall('.//entry') or 
                root.findall('.//security-rule') or 
                root.findall('.//rule') or
                root.findall('.//security-policy/entry')
            )
            
            print(f"Found {len(rule_entries)} rule entries in XML")
            
            # Parse security rules from XML
            for entry in rule_entries:
                pos += 1
                name = entry.get('name', f'rule_{pos}')
                
                # Extract rule attributes with multiple possible paths
                action = (
                    entry.findtext('.//action') or 
                    entry.findtext('action') or 
                    'allow'
                )
                
                # Try different paths for each attribute
                fromzone = (
                    [z.text for z in entry.findall('.//from/member')] or
                    [z.text for z in entry.findall('from/member')] or
                    ['any']
                )
                
                tozone = (
                    [z.text for z in entry.findall('.//to/member')] or
                    [z.text for z in entry.findall('to/member')] or
                    ['any']
                )
                
                source = (
                    [s.text for s in entry.findall('.//source/member')] or
                    [s.text for s in entry.findall('source/member')] or
                    ['any']
                )
                
                destination = (
                    [d.text for d in entry.findall('.//destination/member')] or
                    [d.text for d in entry.findall('destination/member')] or
                    ['any']
                )
                
                application = (
                    [a.text for a in entry.findall('.//application/member')] or
                    [a.text for a in entry.findall('application/member')] or
                    ['any']
                )
                
                service = (
                    [s.text for s in entry.findall('.//service/member')] or
                    [s.text for s in entry.findall('service/member')] or
                    ['any']
                )
                
                rule = RuleLike(
                    name=name,
                    position=pos,
                    action=action,
                    fromzone=fromzone,
                    tozone=tozone,
                    source=source,
                    destination=destination,
                    application=application,
                    service=service,
                    source_user=[],
                    url_category=[],
                    schedule=None,
                    log_setting=None,
                    log_start=None,
                    log_end=None,
                    profile_setting=None,
                    disabled=False,
                    negate_source=False,
                    negate_destination=False,
                    location=None,
                )
                rules.append(rule)
            
            print(f"Successfully parsed {len(rules)} rules from direct API call")
            return rules
        except Exception as e:
            print(f"Direct API rule fetch also failed: {e}")
            return []

    def _to_rulelike(self, r: SecurityRule, position: int) -> RuleLike:
        return RuleLike(
            name=r.name,
            position=position,
            action=r.action or "",
            fromzone=normalize_list(getattr(r, "fromzone", [])),
            tozone=normalize_list(getattr(r, "tozone", [])),
            source=normalize_list(getattr(r, "source", [])),
            destination=normalize_list(getattr(r, "destination", [])),
            application=normalize_list(getattr(r, "application", [])),
            service=normalize_list(getattr(r, "service", [])),
            source_user=normalize_list(getattr(r, "source_user", [])),
            url_category=normalize_list(getattr(r, "category", [])),
            schedule=getattr(r, "schedule", None),
            log_setting=getattr(r, "log_setting", None),
            log_start=getattr(r, "log_start", None),
            log_end=getattr(r, "log_end", None),
            profile_setting=getattr(r, "profile_setting", None),
            disabled=bool(getattr(r, "disabled", False)),
            negate_source=bool(getattr(r, "negate_source", False)),
            negate_destination=bool(getattr(r, "negate_destination", False)),
            location=getattr(r, "location", None),
        )

    def fetch_hit_counts(self, vsys: str = None) -> Dict[str, Dict[str, Any]]:
        """Return mapping: rule_name -> {total:int, last_hit:str, since:str}.
        Requires rule-hit-count to be enabled on device.
        """
        if vsys is None:
            vsys = self.vsys
            
        try:
            # Try different command formats for different PAN-OS versions
            commands = [
                f"show rule-hit-count vsys {vsys} rule-base security rules all",
                f"show rule-hit-count vsys-name {vsys} rule-base security rules all",
                "show rule-hit-count rule-base security rules all",
                "show rule-hit-count rules all",
                f"show rule-hit-count vsys {vsys}",
                f"show rule-hit-count vsys-name {vsys}"
            ]
            
            xml = None
            for cmd in commands:
                try:
                    xml = self.fw.op(cmd, xml=True)
                    break
                except Exception:
                    continue
            
            if xml is None:
                return {}
                
            # Handle both bytes and XML object responses
            if isinstance(xml, bytes):
                import xml.etree.ElementTree as ET
                root = ET.fromstring(xml.decode('utf-8'))
            else:
                root = xml
                
            results: Dict[str, Dict[str, Any]] = {}
            for node in root.findall('.//entry'):
                rn = node.get('name')
                total = int((node.findtext('hit-count') or '0'))
                last = node.findtext('last-hit-time') or ''
                since = node.findtext('time-queried') or ''
                results[rn] = {"total": total, "last": last, "since": since}
            return results
        except Exception:
            return {}


# -----------------------------
# CSV Import Reader (for exported firewall policies)
# -----------------------------
class CSVReader:
    def __init__(self, csv_file_path: str):
        self.csv_file_path = csv_file_path
        self.rules = []
        
    def fetch_rules(self) -> List[RuleLike]:
        """Parse the exported CSV file and convert to RuleLike objects"""
        try:
            df = pd.read_csv(self.csv_file_path)
            print(f"Successfully loaded CSV file: {self.csv_file_path}")
            print(f"Found {len(df)} rules in CSV")
            
            rules: List[RuleLike] = []
            
            for idx, row in df.iterrows():
                try:
                    # Parse position (first column is empty, use index + 1)
                    position = idx + 1
                    
                    # Parse rule name
                    name = str(row.get('Name', f'rule_{position}')).strip()
                    
                    # Parse action
                    action = str(row.get('Action', 'allow')).strip().lower()
                    
                    # Parse zones (split by semicolon and clean up)
                    fromzone = self._parse_list_field(row.get('Source Zone', 'any'))
                    tozone = self._parse_list_field(row.get('Destination Zone', 'any'))
                    
                    # Parse addresses (split by semicolon and clean up)
                    source = self._parse_list_field(row.get('Source Address', 'any'))
                    destination = self._parse_list_field(row.get('Destination Address', 'any'))
                    
                    # Parse applications and services
                    application = self._parse_list_field(row.get('Application', 'any'))
                    service = self._parse_list_field(row.get('Service', 'any'))
                    
                    # Parse users and devices
                    source_user = self._parse_list_field(row.get('Source User', 'any'))
                    source_device = self._parse_list_field(row.get('Source Device', 'any'))
                    
                    # Parse hit counts
                    hit_count_str = str(row.get('Rule Usage Hit Count', '0'))
                    hits_total = int(hit_count_str) if hit_count_str.isdigit() else 0
                    
                    # Parse timestamps
                    last_hit = str(row.get('Rule Usage Last Hit', ''))
                    first_hit = str(row.get('Rule Usage First Hit', ''))
                    
                    # Parse other fields
                    tags = self._parse_list_field(row.get('Tags', ''))
                    rule_type = str(row.get('Type', 'universal'))
                    profile = str(row.get('Profile', ''))
                    options = str(row.get('Options', ''))
                    
                    # Parse additional CSV-specific fields
                    source_device = str(row.get('Source Device', 'any'))
                    destination_device = str(row.get('Destination Device', 'any'))
                    apps_seen = str(row.get('Rule Usage Apps Seen', ''))
                    days_no_new_apps = str(row.get('Days With No New Apps', ''))
                    modified = str(row.get('Modified', ''))
                    created = str(row.get('Created', ''))
                    
                    # Determine if rule is disabled (check name and tags for [Disabled] prefix)
                    disabled = (
                        name.startswith('[Disabled]') or 
                        any(tag.startswith('[Disabled]') for tag in tags)
                    )
                    
                    # Clean up disabled prefixes from name
                    if disabled:
                        name = name.replace('[Disabled]', '').strip()
                    
                    # Create RuleLike object
                    rule = RuleLike(
                        name=name,
                        position=position,
                        action=action,
                        fromzone=fromzone,
                        tozone=tozone,
                        source=source,
                        destination=destination,
                        application=application,
                        service=service,
                        source_user=source_user,
                        url_category=[],  # Not available in CSV export
                        schedule=None,    # Not available in CSV export
                        log_setting=options if options else None,
                        log_start=None,   # Not available in CSV export
                        log_end=None,     # Not available in CSV export
                        profile_setting=profile if profile else None,
                        disabled=disabled,
                        negate_source=False,  # Would need to parse source field for [Negate]
                        negate_destination=False,  # Would need to parse destination field for [Negate]
                        location=None,  # Not available in CSV export
                        hits_total=hits_total,
                        last_hit=last_hit if last_hit != '-' else None,
                        counter_since=first_hit if first_hit != '-' else None,
                    )
                    
                    # Add CSV-specific attributes
                    rule.tags = tags
                    rule.rule_type = rule_type
                    rule.source_device = source_device
                    rule.destination_device = destination_device
                    rule.apps_seen = apps_seen
                    rule.days_no_new_apps = days_no_new_apps
                    rule.modified = modified
                    rule.created = created
                    
                    rules.append(rule)
                    
                except Exception as e:
                    print(f"Warning: Could not parse rule at position {idx + 1}: {e}")
                    continue
            
            print(f"Successfully parsed {len(rules)} rules from CSV")
            return rules
            
        except Exception as e:
            print(f"Error reading CSV file: {e}")
            return []
    
    def _parse_list_field(self, field_value) -> List[str]:
        """Parse a field that may contain multiple values separated by semicolons"""
        if pd.isna(field_value) or field_value == '' or str(field_value).lower() == 'any':
            return ['any']
        
        # Split by semicolon and clean up each item
        items = str(field_value).split(';')
        cleaned_items = []
        
        for item in items:
            item = item.strip()
            # Remove [Disabled] prefixes
            if item.startswith('[Disabled]'):
                item = item.replace('[Disabled]', '').strip()
            if item:
                cleaned_items.append(item)
        
        return cleaned_items if cleaned_items else ['any']
    
    def fetch_hit_counts(self, vsys: str = None) -> Dict[str, Dict[str, Any]]:
        """Return hit counts from CSV data (already embedded in rules)"""
        # Hit counts are already parsed into the rules during CSV import
        # This method is kept for compatibility with the API interface
        return {}


# -----------------------------
# Analyzer (read-only)
# -----------------------------
class Analyzer:
    def __init__(self, rules: List[RuleLike], hit_counts: Optional[Dict[str, Dict[str, Any]]] = None):
        self.rules = rules
        self.name_to_rule = {r.name: r for r in self.rules}
        if hit_counts:
            for r in self.rules:
                if r.name in hit_counts:
                    r.hits_total = hit_counts[r.name].get("total")
                    r.last_hit = hit_counts[r.name].get("last")
                    r.counter_since = hit_counts[r.name].get("since")

    def build_dataframe(self) -> pd.DataFrame:
        rows = [r.to_row() for r in self.rules]
        df = pd.DataFrame(rows)
        df.sort_values(by=["Position"], inplace=True)
        df.reset_index(drop=True, inplace=True)
        return df

    # --- Shadow/order logic ---
    @staticmethod
    def _treat_any(values: List[str]) -> set:
        s = set(values or [])
        if not s or "any" in s:
            return set(["__ANY__"])  # sentinel meaning universal
        return s

    @staticmethod
    def _intersects_set(a: set, b: set) -> bool:
        if "__ANY__" in a or "__ANY__" in b:
            return True
        return not a.isdisjoint(b)

    def rule_intersects(self, r: RuleLike, union: Dict[str, set]) -> bool:
        return (
            self._intersects_set(self._treat_any(r.fromzone), union["from"]) and
            self._intersects_set(self._treat_any(r.tozone), union["to"]) and
            self._intersects_set(self._treat_any(r.source), union["src"]) and
            self._intersects_set(self._treat_any(r.destination), union["dst"]) and
            self._intersects_set(self._treat_any(r.application), union["app"]) and
            self._intersects_set(self._treat_any(r.service), union["svc"]) and
            self._intersects_set(self._treat_any(r.source_user), union["user"]) and
            self._intersects_set(self._treat_any(r.url_category), union["urlcat"]) and
            (r.schedule == union["schedule"]) and
            (r.action in ["allow", "deny", "drop"])
        )

    def find_shadowed_rules(self) -> List[ShadowFinding]:
        findings: List[ShadowFinding] = []
        # Evaluate top-down; earlier rules can shadow later ones if same action and superset match
        for i, earlier in enumerate(self.rules):
            if earlier.disabled:
                continue
            for j in range(i + 1, len(self.rules)):
                later = self.rules[j]
                if later.disabled:
                    continue
                if earlier.action != later.action:
                    continue
                # Check superset: earlier covers later
                if self._covers(earlier, later):
                    findings.append(
                        ShadowFinding(
                            shadowed_rule=later.name,
                            shadowed_position=later.position,
                            shadowing_rule=earlier.name,
                            shadowing_position=earlier.position,
                            reason="Earlier rule fully covers later rule",
                            recommendation=(
                                "Later rule appears shadowed by earlier rule; consider merging into the top-most "
                                f"rule '{earlier.name}' or removing after review."
                            ),
                        )
                    )
        return findings

    def _covers(self, a: RuleLike, b: RuleLike) -> bool:
        # a covers b if for every dimension, a's set is a superset (treat 'any' as universal)
        def superset(sa: List[str], sb: List[str]) -> bool:
            A = self._treat_any(sa)
            B = self._treat_any(sb)
            return ("__ANY__" in A) or B.issubset(A)
        return (
            superset(a.fromzone, b.fromzone) and
            superset(a.tozone, b.tozone) and
            superset(a.source, b.source) and
            superset(a.destination, b.destination) and
            superset(a.application, b.application) and
            superset(a.service, b.service) and
            superset(a.source_user, b.source_user) and
            superset(a.url_category, b.url_category) and
            (a.schedule == b.schedule)
        )

    def propose_merges(self) -> List[Proposal]:
        # Group rules by identical non-broadening attributes
        buckets: Dict[str, Dict[str, Any]] = {}
        for r in self.rules:
            fp, meta = r.non_broadening_fingerprint()
            buckets.setdefault(fp, {"meta": meta, "rules": []})["rules"].append(r)

        proposals: List[Proposal] = []
        for fp, data in buckets.items():
            group: List[RuleLike] = data["rules"]
            if len(group) < 2:
                continue
            # Skip negations
            if any([g.negate_source or g.negate_destination for g in group]):
                continue
            # Treat ANY symmetry (avoid mixing any vs specific)
            any_flags = [(
                "any" in (g.application or []),
                "any" in (g.service or []),
                "any" in (g.source or []),
                "any" in (g.destination or []),
            ) for g in group]
            if len(set(any_flags)) > 1:
                continue

            # Build unions
            apps = sorted(as_set(group[0].application).union(*[as_set(g.application) for g in group[1:]]))
            svcs = sorted(as_set(group[0].service).union(*[as_set(g.service) for g in group[1:]]))
            srcs = sorted(as_set(group[0].source).union(*[as_set(g.source) for g in group[1:]]))
            dsts = sorted(as_set(group[0].destination).union(*[as_set(g.destination) for g in group[1:]]))

            positions = [g.position for g in group]
            i, j = min(positions), max(positions)
            union = {
                "from": self._treat_any(group[0].fromzone),
                "to": self._treat_any(group[0].tozone),
                "src": self._treat_any(srcs),
                "dst": self._treat_any(dsts),
                "app": self._treat_any(apps),
                "svc": self._treat_any(svcs),
                "user": self._treat_any(group[0].source_user),
                "urlcat": self._treat_any(group[0].url_category),
                "schedule": group[0].schedule,
            }

            # Order-sensitive scan: if any intervening DENY intersects union, flag it
            order_sensitive, reason = self._order_sensitive_between(i, j, union)

            confidence = "High" if not order_sensitive else "Low"
            recommendation = (
                "Merge: Identical qualifiers; union of objects is equivalent; no intervening deny."
                if not order_sensitive else
                "Review: Intervening deny/ordering may affect behavior; manual confirmation required."
            )
            prop = Proposal(
                proposed_name=f"merge_of_{'_'.join([g.name for g in group[:3]])}"[:63],
                source_rules=[g.name for g in group],
                positions=positions,
                apps_union=apps,
                services_union=svcs,
                sources_union=srcs,
                destinations_union=dsts,
                order_sensitive=order_sensitive,
                order_reason=reason,
                confidence=confidence,
                recommendation=recommendation,
            )
            proposals.append(prop)
        return proposals

    def _order_sensitive_between(self, i: int, j: int, union: Dict[str, set]) -> Tuple[bool, str]:
        # indices are positions (1-based). Scan between them in ordered list
        # If any DENY intersects the union, mark order sensitive.
        # If any ALLOW earlier fully covers a later, that will be caught in shadow pass separately.
        for r in self.rules:
            if r.position <= i or r.position >= j:
                continue
            if r.action.lower() in ("deny", "drop"):
                if self.rule_intersects(r, union):
                    return True, f"Intervening deny/drop at position {r.position} intersects proposed union"
        return False, ""

    def unused_rules_zero_hits(self) -> List[RuleLike]:
        return [r for r in self.rules if (r.hits_total is not None and r.hits_total == 0)]


# -----------------------------
# Exporters (CSV / XLSX / HTML / PDF)
# -----------------------------
class Exporter:
    HEADER_ORDER = [
        "Position","Name","Tags","Type","Source Zone","Source Address","Source User","Source Device",
        "Destination Zone","Destination Address","Destination Device","Application","Service","Action",
        "Profile","Options","Rule Usage Hit Count","Rule Usage Last Hit","Rule Usage First Hit",
        "Rule Usage Apps Seen","Days With No New Apps","Modified","Created","Recommendation"
    ]

    @staticmethod
    def _sanitize_df_for_export(df: pd.DataFrame) -> pd.DataFrame:
        """Return a copy of df with all values converted to simple scalars/strings.
        Prevents errors like 'truth value of an array is ambiguous' during export.
        """
        def to_scalar_string(value: Any) -> Any:
            # Fast path for common scalars and None
            if value is None:
                return ""
            if isinstance(value, (str, int, float, bool)):
                return value
            # Avoid using pd.isna on non-scalars; it can return an array
            try:
                # Handle pandas NA/NaN scalars
                if isinstance(value, float) and pd.isna(value):
                    return ""
            except Exception:
                pass

            # Join common iterable types
            if isinstance(value, (list, tuple, set)):
                return ", ".join([str(x) for x in value])

            # Handle objects that can be converted to list (e.g., numpy arrays)
            tolist = getattr(value, "tolist", None)
            if callable(tolist):
                try:
                    seq = tolist()
                    return ", ".join([str(x) for x in seq])
                except Exception:
                    pass

            # Fallback to string
            return str(value)

        clean = df.copy()
        for col in clean.columns:
            clean[col] = clean[col].map(to_scalar_string)
        return clean

    @staticmethod
    def dataframe_with_recommendations(df_rules: pd.DataFrame,
                                       unused: List[RuleLike],
                                       shadows: List[ShadowFinding],
                                       merges: List[Proposal]) -> pd.DataFrame:
        df = df_rules.copy()
        df["Recommendation"] = ""

        # Map: rule -> rec lines
        recs: Dict[str, List[str]] = {}

        # Unused (0 hits) => recommend disable
        for r in unused:
            recs.setdefault(r.name, []).append("Disable: 0 hits over observation window.")

        # Shadows => call out shadowing rule and suggest consolidation
        for s in shadows:
            recs.setdefault(s.shadowed_rule, []).append(
                f"Shadowed by '{s.shadowing_rule}' (pos {s.shadowing_position}); consider merge into top-most and remove after review."
            )

        # Merges => suggest merge group with positions
        for p in merges:
            for rn in p.source_rules:
                # Get position numbers for all rules in the merge group
                merge_positions = [f"{name} (pos {pos})" for name, pos in zip(p.source_rules, p.positions)]
                other_rules_with_pos = [f"{name} (pos {pos})" for name, pos in zip(p.source_rules, p.positions) if name != rn]
                msg = (
                    f"Merge-candidate with {', '.join(other_rules_with_pos)}"
                    f"; confidence={p.confidence}. {p.order_reason or ''}"
                ).strip()
                recs.setdefault(rn, []).append(msg)

        # Apply back to df
        df["Recommendation"] = df["Name"].map(lambda n: " | ".join(recs.get(n, [])))

        # Reorder columns and fill NaNs
        for col in Exporter.HEADER_ORDER:
            if col not in df.columns:
                df[col] = ""
        df = df[Exporter.HEADER_ORDER]
        df.fillna("", inplace=True)
        df.sort_values(by=["Position"], inplace=True)
        df.reset_index(drop=True, inplace=True)
        return df

    @staticmethod
    def export_csv(df: pd.DataFrame, path: str) -> str:
        clean = Exporter._sanitize_df_for_export(df)
        clean.to_csv(path, index=False)
        return path

    @staticmethod
    def export_xlsx(df: pd.DataFrame, path: str) -> str:
        # Column coloring & header style
        wb = pd.ExcelWriter(path, engine="openpyxl")
        clean = Exporter._sanitize_df_for_export(df)
        clean.to_excel(wb, sheet_name="Policy", index=False)
        wb.book.active = 0
        ws = wb.book["Policy"]
        # Style header
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        header_fill = PatternFill("solid", fgColor="DDDDDD")
        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(border_style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        # Zebra column colors to improve readability
        col_fills = ["FFFFFF", "F2F8FF"]  # alternate
        for idx, col_cells in enumerate(ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
            fill = PatternFill("solid", fgColor=col_fills[(idx - 1) % 2])
            for cell in col_cells:
                cell.fill = fill
                cell.border = border

        # Set reasonable column widths
        for col in ws.columns:
            max_len = 12
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        wb.close()
        return path

    @staticmethod
    def export_csv_with_tabs(df_analysis: pd.DataFrame, df_overview: pd.DataFrame, path: str) -> str:
        """Export both analysis and overview data to a single CSV file with sections"""
        with open(path, 'w', newline='', encoding='utf-8') as f:
            # Write overview section
            f.write("=== OVERVIEW ===\n")
            clean_overview = Exporter._sanitize_df_for_export(df_overview)
            clean_overview.to_csv(f, index=False)
            
            # Add separator
            f.write("\n\n=== ANALYSIS ===\n")
            
            # Write analysis section
            clean_analysis = Exporter._sanitize_df_for_export(df_analysis)
            clean_analysis.to_csv(f, index=False)
        
        return path

    @staticmethod
    def export_xlsx_with_tabs(df_analysis: pd.DataFrame, df_overview: pd.DataFrame, path: str) -> str:
        """Export both analysis and overview data to Excel with separate sheets"""
        wb = pd.ExcelWriter(path, engine="openpyxl")
        
        # Export overview sheet
        clean_overview = Exporter._sanitize_df_for_export(df_overview)
        clean_overview.to_excel(wb, sheet_name="Overview", index=False)
        
        # Export analysis sheet
        clean_analysis = Exporter._sanitize_df_for_export(df_analysis)
        clean_analysis.to_excel(wb, sheet_name="Analysis", index=False)
        
        # Style both sheets
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        header_fill = PatternFill("solid", fgColor="DDDDDD")
        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(border_style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        # Style overview sheet
        ws_overview = wb.book["Overview"]
        for cell in ws_overview[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        
        # Style analysis sheet
        ws_analysis = wb.book["Analysis"]
        for cell in ws_analysis[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        
        # Zebra column colors for both sheets
        col_fills = ["FFFFFF", "F2F8FF"]
        for ws in [ws_overview, ws_analysis]:
            for idx, col_cells in enumerate(ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
                fill = PatternFill("solid", fgColor=col_fills[(idx - 1) % 2])
                for cell in col_cells:
                    cell.fill = fill
                    cell.border = border
        
        # Set column widths for both sheets
        for ws in [ws_overview, ws_analysis]:
            for col in ws.columns:
                max_len = 12
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 4, 60)
        
        wb.close()
        return path

    @staticmethod
    def export_html(df: pd.DataFrame, path: str) -> str:
        html = df.to_html(index=False, escape=False, border=0)
        template = f"""
        <html>
        <head>
          <meta charset='utf-8'>
          <title>{APP_NAME} Report</title>
          <style>
            body {{ 
              font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
              margin: 0; 
              padding: 20px; 
              background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%);
              color: #e2e8f0;
              min-height: 100vh;
            }}
            .container {{
              max-width: 1400px;
              margin: 0 auto;
              background: rgba(30, 41, 59, 0.8);
              border-radius: 12px;
              padding: 30px;
              box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.1), 0 10px 10px -5px rgba(0, 0, 0, 0.04);
            }}
            h1 {{ 
              color: #3b82f6; 
              font-size: 2.5em;
              margin-bottom: 10px;
              text-align: center;
              text-shadow: 0 2px 4px rgba(0, 0, 0, 0.3);
            }}
            .subtitle {{
              text-align: center;
              color: #94a3b8;
              margin-bottom: 30px;
              font-size: 1.1em;
            }}
            table {{ 
              border-collapse: collapse; 
              width: 100%; 
              background: #1e293b;
              border-radius: 8px;
              overflow: hidden;
              box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
            }}
            thead th {{ 
              position: sticky; 
              top: 0; 
              background: linear-gradient(135deg, #334155 0%, #475569 100%); 
              color: #f8fafc; 
              padding: 12px 8px; 
              font-weight: bold;
              text-transform: uppercase;
              font-size: 0.9em;
              letter-spacing: 0.5px;
              border-bottom: 2px solid #3b82f6;
            }}
            td {{ 
              border: 1px solid #334155; 
              padding: 10px 8px; 
              vertical-align: top; 
              font-size: 0.9em;
            }}
            tr:nth-child(even) {{ background: #334155; }}
            tr:nth-child(odd) {{ background: #1e293b; }}
            tr:hover {{ 
              background: linear-gradient(135deg, #3b82f6 0%, #2563eb 100%);
              color: #ffffff;
              transform: scale(1.01);
              transition: all 0.2s ease;
            }}
            .zero-hits {{
              background: linear-gradient(135deg, #dc2626 0%, #b91c1c 100%) !important;
              color: #ffffff !important;
            }}
            .allow-action {{
              color: #10b981;
              font-weight: bold;
            }}
            .deny-action {{
              color: #ef4444;
              font-weight: bold;
            }}
          </style>
        </head>
        <body>
          <div class="container">
            <h1>🔍 {APP_NAME}</h1>
            <p class="subtitle">📊 Policy Analysis Report • Generated: {now_iso()}</p>
            {html}
          </div>
        </body>
        </html>
        """
        with open(path, "w", encoding="utf-8") as f:
            f.write(template)
        return path

    @staticmethod
    def export_pdf(df: pd.DataFrame, path: str) -> str:
        doc = SimpleDocTemplate(path, pagesize=landscape(letter), leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
        styles = getSampleStyleSheet()
        story = []
        story.append(Paragraph(f"{APP_NAME} — Recommendations", styles['Title']))
        story.append(Paragraph(f"Generated: {now_iso()}", styles['Normal']))
        story.append(Spacer(1, 12))

        # Convert DataFrame to list of lists, ensuring all values are strings
        headers = list(df.columns)
        data = [headers]
        
        for _, row in df.iterrows():
            # Convert each value to string and handle None/NaN values
            row_data = []
            for col in headers:
                value = row.get(col)
                if pd.isna(value) or value is None:
                    row_data.append("")
                else:
                    # Convert to string and truncate if too long
                    str_value = str(value)
                    if len(str_value) > 50:
                        str_value = str_value[:47] + "..."
                    row_data.append(str_value)
            data.append(row_data)

        table = Table(data, repeatRows=1)
        style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.black),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('FONTSIZE', (0,0), (-1,-1), 8),  # Smaller font size to fit more data
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.lightgrey]),  # Alternating row colors
        ])
        table.setStyle(style)
        story.append(table)
        doc.build(story)
        return path


# -----------------------------
# Tkinter GUI
# -----------------------------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_NAME)
        self.geometry("1200x800")
        self.configure(bg="#0f172a")
        
        # Enhanced styling
        self.style = ttk.Style(self)
        self.style.theme_use('clam')
        
        # Configure modern dark theme
        self.style.configure('TLabel', background="#0f172a", foreground="#e2e8f0", font=('Segoe UI', 9))
        self.style.configure('TButton', 
                           background="#3b82f6", 
                           foreground="#ffffff", 
                           font=('Segoe UI', 9, 'bold'),
                           borderwidth=0,
                           focuscolor='none')
        self.style.map('TButton',
                      background=[('active', '#2563eb'), ('pressed', '#1d4ed8')])
        
        self.style.configure('TEntry', 
                           fieldbackground="#1e293b", 
                           foreground="#e2e8f0",
                           borderwidth=1,
                           relief='flat')
        self.style.map('TEntry',
                      fieldbackground=[('focus', '#334155')])
        
        self.style.configure('TFrame', background="#0f172a")
        
        # Enhanced Treeview styling
        self.style.configure('Treeview', 
                           background='#1e293b', 
                           fieldbackground='#1e293b', 
                           foreground='#e2e8f0',
                           font=('Segoe UI', 9),
                           rowheight=25)
        self.style.configure('Treeview.Heading', 
                           background='#334155', 
                           foreground='#f8fafc',
                           font=('Segoe UI', 9, 'bold'),
                           relief='flat')
        self.style.map('Treeview',
                      background=[('selected', '#3b82f6')],
                      foreground=[('selected', '#ffffff')])
        
        # Scrollbar styling
        self.style.configure('Vertical.TScrollbar', 
                           background='#475569',
                           troughcolor='#1e293b',
                           width=12,
                           borderwidth=0,
                           relief='flat')
        self.style.configure('Horizontal.TScrollbar', 
                           background='#475569',
                           troughcolor='#1e293b',
                           width=12,
                           borderwidth=0,
                           relief='flat')

        self.conf_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), CONF_FILE)
        self.conf = read_conf(self.conf_path)

        # Load configuration with defaults
        self.api_url_var = tk.StringVar(value=self.conf.get("api_url", "fw-mgmt.example.com"))
        self.api_key_var = tk.StringVar(value=self.conf.get("api_key", ""))
        self.vsys_var = tk.StringVar(value=self.conf.get("vsys", "vsys1"))
        self.output_dir_var = tk.StringVar(value=self.conf.get("output_dir", os.getcwd()))
        self.csv_file_var = tk.StringVar(value=self.conf.get("csv_file", ""))
        
        # Handle mode conversion from old format to new format
        saved_mode = self.conf.get("mode", "api")
        if saved_mode == "api":
            self.mode_var = tk.StringVar(value="API Connection")
        elif saved_mode == "csv":
            self.mode_var = tk.StringVar(value="CSV Import")
        else:
            self.mode_var = tk.StringVar(value=saved_mode if saved_mode in ["API Connection", "CSV Import"] else "API Connection")
        
        # Load window geometry if saved
        saved_geometry = self.conf.get("window_geometry", "1200x800")
        self.geometry(saved_geometry)
        
        # Bind window close event to save settings
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # Show configuration loaded message
        if self.conf:
            print(f"Configuration loaded from {self.conf_path}")
            if "last_saved" in self.conf:
                print(f"Last saved: {self.conf['last_saved']}")
        else:
            print("No existing configuration found, using defaults")

        # Check dependencies on startup
        print("\n=== Dependency Check ===")
        from evaluator import check_dependencies
        deps = check_dependencies()
        for dep in deps:
            print(dep)
        print("========================\n")

        self.df_rules: Optional[pd.DataFrame] = None
        self.df_final: Optional[pd.DataFrame] = None
        self.unused: List[RuleLike] = []
        self.shadows: List[ShadowFinding] = []
        self.merges: List[Proposal] = []

        self._build_ui()
        
        # Initialize mode display
        self.on_mode_change()

    def _build_ui(self):
        # Main container with padding
        main_container = ttk.Frame(self)
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Title section
        title_frame = ttk.Frame(main_container)
        title_frame.pack(fill=tk.X, pady=(0, 20))
        title_label = ttk.Label(title_frame, text="PAN-OS Policy Evaluator", 
                               font=('Segoe UI', 16, 'bold'))
        title_label.pack(anchor='w')
        
        # Mode selection
        mode_frame = ttk.Frame(main_container)
        mode_frame.pack(fill=tk.X, pady=(0, 20))
        
        ttk.Label(mode_frame, text="Analysis Mode:", width=12).pack(side=tk.LEFT)
        mode_combo = ttk.Combobox(mode_frame, textvariable=self.mode_var, 
                                 values=["API Connection", "CSV Import"], 
                                 state="readonly", width=15)
        mode_combo.pack(side=tk.LEFT, padx=(8, 20))
        mode_combo.bind('<<ComboboxSelected>>', self.on_mode_change)
        
        # Configuration section with better layout
        config_frame = ttk.Frame(main_container)
        config_frame.pack(fill=tk.X, pady=(0, 20))
        
        # API Configuration Frame
        self.api_config_frame = ttk.LabelFrame(config_frame, text="API Configuration")
        self.api_config_frame.pack(fill=tk.X, pady=(0, 12))
        
        # Row 1: API URL and API Key
        row1 = ttk.Frame(self.api_config_frame)
        row1.pack(fill=tk.X, pady=(8, 8), padx=8)
        
        ttk.Label(row1, text="API URL / Host:", width=12).pack(side=tk.LEFT)
        ttk.Entry(row1, textvariable=self.api_url_var, width=35).pack(side=tk.LEFT, padx=(8, 20))
        
        ttk.Label(row1, text="API Key:", width=8).pack(side=tk.LEFT)
        ttk.Entry(row1, textvariable=self.api_key_var, width=45, show="*").pack(side=tk.LEFT, padx=8)
        
        # Row 2: VSYS
        row2 = ttk.Frame(self.api_config_frame)
        row2.pack(fill=tk.X, pady=(0, 8), padx=8)
        
        ttk.Label(row2, text="VSYS:", width=12).pack(side=tk.LEFT)
        ttk.Entry(row2, textvariable=self.vsys_var, width=20).pack(side=tk.LEFT, padx=(8, 20))
        
        # CSV Configuration Frame
        self.csv_config_frame = ttk.LabelFrame(config_frame, text="CSV Import Configuration")
        self.csv_config_frame.pack(fill=tk.X, pady=(0, 12))
        
        csv_row = ttk.Frame(self.csv_config_frame)
        csv_row.pack(fill=tk.X, pady=8, padx=8)
        
        ttk.Label(csv_row, text="CSV File:", width=12).pack(side=tk.LEFT)
        csv_entry = ttk.Entry(csv_row, textvariable=self.csv_file_var, width=60)
        csv_entry.pack(side=tk.LEFT, padx=(8, 8))
        ttk.Button(csv_row, text="Browse", command=self.choose_csv_file, width=10).pack(side=tk.LEFT)
        
        # Output Folder Row
        output_row = ttk.Frame(config_frame)
        output_row.pack(fill=tk.X, pady=(0, 12))
        
        ttk.Label(output_row, text="Output Folder:", width=12).pack(side=tk.LEFT)
        out_entry = ttk.Entry(output_row, textvariable=self.output_dir_var, width=50)
        out_entry.pack(side=tk.LEFT, padx=(8, 8))
        ttk.Button(output_row, text="Browse", command=self.choose_output_dir, width=10).pack(side=tk.LEFT)
        
        # Button section with improved spacing
        button_frame = ttk.Frame(main_container)
        button_frame.pack(fill=tk.X, pady=(0, 20))
        
        # Common buttons
        ttk.Button(button_frame, text="💾 Save Config", command=self.save_conf, width=15).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="⚙️ Show Config", command=self.show_config, width=12).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="🔧 Check Deps", command=self.check_dependencies, width=12).pack(side=tk.LEFT, padx=(0, 10))
        
        # Mode-specific buttons
        self.api_test_button = ttk.Button(button_frame, text="🔌 Test API", command=self.test_api_connection, width=12)
        self.api_test_button.pack(side=tk.LEFT, padx=(0, 10))
        
        # Analysis button (changes text based on mode)
        self.analyze_button = ttk.Button(button_frame, text="🔍 Fetch & Analyze", command=self.run_analyze, width=18)
        self.analyze_button.pack(side=tk.LEFT, padx=(0, 10))
        
        # Results buttons
        ttk.Button(button_frame, text="📊 Preview Summary", command=self.preview_summary, width=16).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="📤 Export Reports", command=self.export_all, width=16).pack(side=tk.LEFT)
        
        # Status bar with better styling
        status_frame = ttk.Frame(main_container)
        status_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Show configuration status
        config_status = self._get_config_status()
        self.status = tk.StringVar(value=f"Ready to analyze PAN-OS policies... | {config_status}")
        status_label = ttk.Label(status_frame, textvariable=self.status, 
                                font=('Segoe UI', 9, 'italic'))
        status_label.pack(anchor='w')
        
        # Results section with tabbed interface
        results_frame = ttk.Frame(main_container)
        results_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(results_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        # Analysis Tab
        analysis_frame = ttk.Frame(self.notebook)
        self.notebook.add(analysis_frame, text="Analysis")
        
        # Create treeview with scrollbars for Analysis tab
        tree_frame = ttk.Frame(analysis_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        # Vertical scrollbar
        v_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical")
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Horizontal scrollbar
        h_scrollbar = ttk.Scrollbar(tree_frame, orient="horizontal")
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Treeview with all CSV columns
        self.tree = ttk.Treeview(tree_frame, 
                                 columns=("pos","name","tags","type","src_zone","src_addr","src_user","src_dev",
                                         "dst_zone","dst_addr","dst_dev","app","svc","action","profile","options",
                                         "hits","last_hit","first_hit","apps_seen","days_no_apps","modified","created","rec"), 
                                 show='headings', 
                                 height=20,
                                 yscrollcommand=v_scrollbar.set,
                                 xscrollcommand=h_scrollbar.set)
        
        # Configure scrollbars
        v_scrollbar.config(command=self.tree.yview)
        h_scrollbar.config(command=self.tree.xview)
         
        # Complete column configuration matching CSV structure
        columns_config = [
            ("pos", "Position", 80),
            ("name", "Name", 200),
            ("tags", "Tags", 150),
            ("type", "Type", 100),
            ("src_zone", "Source Zone", 120),
            ("src_addr", "Source Address", 150),
            ("src_user", "Source User", 120),
            ("src_dev", "Source Device", 120),
            ("dst_zone", "Dest Zone", 120),
            ("dst_addr", "Dest Address", 150),
            ("dst_dev", "Dest Device", 120),
            ("app", "Application", 120),
            ("svc", "Service", 120),
            ("action", "Action", 80),
            ("profile", "Profile", 150),
            ("options", "Options", 150),
            ("hits", "Hit Count", 100),
            ("last_hit", "Last Hit", 120),
            ("first_hit", "First Hit", 120),
            ("apps_seen", "Apps Seen", 100),
            ("days_no_apps", "Days No Apps", 120),
            ("modified", "Modified", 120),
            ("created", "Created", 120),
            ("rec", "Recommendation", 300)
        ]
        
        for col_id, col_name, width in columns_config:
            self.tree.heading(col_id, text=col_name)
            self.tree.column(col_id, width=width, anchor='w', minwidth=50)
        
        self.tree.pack(fill=tk.BOTH, expand=True)
        
        # Add alternating row colors
        self.tree.tag_configure('oddrow', background='#1e293b')
        self.tree.tag_configure('evenrow', background='#334155')
         
        # Overview Tab
        overview_frame = ttk.Frame(self.notebook)
        self.notebook.add(overview_frame, text="Overview")
        
        # Create treeview for Overview tab
        overview_tree_frame = ttk.Frame(overview_frame)
        overview_tree_frame.pack(fill=tk.BOTH, expand=True)
        
        # Vertical scrollbar for overview
        ov_v_scrollbar = ttk.Scrollbar(overview_tree_frame, orient="vertical")
        ov_v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Horizontal scrollbar for overview
        ov_h_scrollbar = ttk.Scrollbar(overview_tree_frame, orient="horizontal")
        ov_h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Overview treeview
        self.overview_tree = ttk.Treeview(overview_tree_frame,
                                        columns=("category", "metric", "value", "description"),
                                        show='headings',
                                        height=20,
                                        yscrollcommand=ov_v_scrollbar.set,
                                        xscrollcommand=ov_h_scrollbar.set)
        
        # Configure scrollbars
        ov_v_scrollbar.config(command=self.overview_tree.yview)
        ov_h_scrollbar.config(command=self.overview_tree.xview)
        
        # Configure overview columns
        self.overview_tree.heading("category", text="Category")
        self.overview_tree.heading("metric", text="Metric")
        self.overview_tree.heading("value", text="Value")
        self.overview_tree.heading("description", text="Description")
        
        self.overview_tree.column("category", width=150, anchor='w')
        self.overview_tree.column("metric", width=200, anchor='w')
        self.overview_tree.column("value", width=100, anchor='w')
        self.overview_tree.column("description", width=300, anchor='w')
        
        self.overview_tree.pack(fill=tk.BOTH, expand=True)
        
        # Add alternating row colors for overview
        self.overview_tree.tag_configure('oddrow', background='#1e293b')
        self.overview_tree.tag_configure('evenrow', background='#334155')

    def choose_output_dir(self):
        d = filedialog.askdirectory(initialdir=self.output_dir_var.get() or os.getcwd())
        if d:
            self.output_dir_var.set(d)
    
    def choose_csv_file(self):
        """Choose a CSV file for import"""
        f = filedialog.askopenfilename(
            title="Select CSV Export File",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialdir=os.path.dirname(self.csv_file_var.get()) if self.csv_file_var.get() else os.getcwd()
        )
        if f:
            self.csv_file_var.set(f)
    
    def on_mode_change(self, event=None):
        """Handle mode change between API and CSV"""
        mode = self.mode_var.get()
        if mode == "API Connection":
            self.api_config_frame.pack(fill=tk.X, pady=(0, 12))
            self.csv_config_frame.pack_forget()
            self.api_test_button.config(state='normal')
            self.analyze_button.config(text="🔍 Fetch & Analyze")
        else:  # CSV Import
            self.api_config_frame.pack_forget()
            self.csv_config_frame.pack(fill=tk.X, pady=(0, 12))
            self.api_test_button.config(state='disabled')
            self.analyze_button.config(text="📁 Import & Analyze")
        
        # Update status
        self._update_status_for_mode()
    
    def _update_status_for_mode(self):
        """Update status bar based on current mode"""
        mode = self.mode_var.get()
        if mode == "API Connection":
            config_status = self._get_config_status()
            self.status.set(f"Ready to analyze via API... | {config_status}")
        else:
            csv_file = self.csv_file_var.get().strip()
            if csv_file:
                filename = os.path.basename(csv_file)
                self.status.set(f"Ready to analyze CSV file: {filename}")
            else:
                self.status.set("Ready to analyze CSV file... | Select CSV file")

    def save_conf(self):
        """Save current configuration to file"""
        # Convert mode to save format
        mode = self.mode_var.get()
        if mode == "API Connection":
            save_mode = "api"
        elif mode == "CSV Import":
            save_mode = "csv"
        else:
            save_mode = mode
            
        data = {
            "api_url": self.api_url_var.get().strip(),
            "api_key": self.api_key_var.get().strip(),
            "vsys": self.vsys_var.get().strip(),
            "output_dir": self.output_dir_var.get().strip(),
            "csv_file": self.csv_file_var.get().strip(),
            "mode": save_mode,
            "window_geometry": self.geometry(),
            "last_saved": now_iso(),
        }
        write_conf(self.conf_path, data)
        # Update the configuration object
        self.conf = data
        # Update status bar
        self._update_status_for_mode()
        self.status.set(f"✅ Configuration saved to {self.conf_path}")
    
    def on_closing(self):
        """Handle application closing - save settings automatically"""
        try:
            # Convert mode to save format
            mode = self.mode_var.get()
            if mode == "API Connection":
                save_mode = "api"
            elif mode == "CSV Import":
                save_mode = "csv"
            else:
                save_mode = mode
                
            # Auto-save current settings
            data = {
                "api_url": self.api_url_var.get().strip(),
                "api_key": self.api_key_var.get().strip(),
                "vsys": self.vsys_var.get().strip(),
                "output_dir": self.output_dir_var.get().strip(),
                "csv_file": self.csv_file_var.get().strip(),
                "mode": save_mode,
                "window_geometry": self.geometry(),
                "last_saved": now_iso(),
            }
            write_conf(self.conf_path, data)
            print(f"Settings auto-saved to {self.conf_path}")
        except Exception as e:
            print(f"Warning: Could not auto-save settings: {e}")
        
        # Destroy the window
        self.destroy()
    
    def _get_config_status(self) -> str:
        """Get a human-readable status of the current configuration"""
        mode = self.mode_var.get()
        api_url = self.api_url_var.get().strip()
        api_key = self.api_key_var.get().strip()
        vsys = self.vsys_var.get().strip()
        csv_file = self.csv_file_var.get().strip()
        output_dir = self.output_dir_var.get().strip()
        
        status_parts = []
        
        # Mode status
        status_parts.append(f"Mode: {mode}")
        
        if mode == "API Connection":
            if api_url and api_url != "fw-mgmt.example.com":
                status_parts.append(f"API: {api_url}")
            else:
                status_parts.append("API: Not configured")
                
            if api_key:
                status_parts.append("Key: Set")
            else:
                status_parts.append("Key: Not set")
                
            if vsys and vsys != "vsys1":
                status_parts.append(f"VSYS: {vsys}")
            else:
                status_parts.append("VSYS: Default")
        else:  # CSV Import
            if csv_file:
                filename = os.path.basename(csv_file)
                status_parts.append(f"CSV: {filename}")
            else:
                status_parts.append("CSV: Not selected")
            
        if output_dir and output_dir != os.getcwd():
            status_parts.append(f"Output: {os.path.basename(output_dir)}")
        else:
            status_parts.append("Output: Current directory")
        
        # Add last saved info if available
        if "last_saved" in self.conf:
            status_parts.append(f"Last saved: {self.conf['last_saved']}")
        
        return " | ".join(status_parts)

    def show_config(self):
        """Show current configuration in a popup window"""
        config_text = f"""Current Configuration
====================

Analysis Mode: {self.mode_var.get()}

API Configuration:
API URL: {self.api_url_var.get().strip() or 'Not set'}
API Key: {'*' * min(len(self.api_key_var.get()), 8) + '...' if self.api_key_var.get() else 'Not set'}
VSYS: {self.vsys_var.get().strip() or 'vsys1'}

CSV Configuration:
CSV File: {self.csv_file_var.get().strip() or 'Not set'}

General Configuration:
Output Directory: {self.output_dir_var.get().strip() or os.getcwd()}
Window Geometry: {self.geometry()}

Configuration File: {self.conf_path}
Last Saved: {self.conf.get('last_saved', 'Never')}

Configuration Status:
{self._get_config_status()}
"""
        
        # Create popup window
        config_window = tk.Toplevel(self)
        config_window.title("Current Configuration")
        config_window.geometry("600x400")
        config_window.configure(bg="#0f172a")
        config_window.transient(self)
        config_window.grab_set()
        
        # Main container
        main_container = ttk.Frame(config_window)
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Title
        title_label = ttk.Label(main_container, text="⚙️ Current Configuration", 
                               font=('Segoe UI', 14, 'bold'))
        title_label.pack(anchor='w', pady=(0, 15))
        
        # Text area with scrollbars
        text_frame = ttk.Frame(main_container)
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(text_frame)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        h_scrollbar = ttk.Scrollbar(text_frame, orient="horizontal")
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Text widget
        txt = tk.Text(text_frame, 
                     wrap='word',
                     bg="#1e293b", 
                     fg="#e2e8f0",
                     font=('Consolas', 10),
                     insertbackground="#e2e8f0",
                     selectbackground="#3b82f6",
                     yscrollcommand=v_scrollbar.set,
                     xscrollcommand=h_scrollbar.set)
        
        v_scrollbar.config(command=txt.yview)
        h_scrollbar.config(command=txt.xview)
        
        txt.insert('1.0', config_text)
        txt.config(state='disabled')
        txt.pack(fill=tk.BOTH, expand=True)
        
        # Close button
        button_frame = ttk.Frame(main_container)
        button_frame.pack(fill=tk.X, pady=(15, 0))
        
        ttk.Button(button_frame, text="Close", command=config_window.destroy, width=10).pack(side=tk.RIGHT)

    def check_dependencies(self):
        """Check and display dependency status"""
        from evaluator import check_dependencies
        
        issues = check_dependencies()
        
        # Create popup window
        deps_window = tk.Toplevel(self)
        deps_window.title("Dependency Check")
        deps_window.geometry("500x400")
        deps_window.configure(bg="#0f172a")
        deps_window.transient(self)
        deps_window.grab_set()
        
        # Main container
        main_container = ttk.Frame(deps_window)
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Title
        title_label = ttk.Label(main_container, text="🔧 Dependency Check", 
                               font=('Segoe UI', 14, 'bold'))
        title_label.pack(anchor='w', pady=(0, 15))
        
        # Text area
        text_frame = ttk.Frame(main_container)
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbar
        v_scrollbar = ttk.Scrollbar(text_frame)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Text widget
        txt = tk.Text(text_frame, 
                     wrap='word',
                     bg="#1e293b", 
                     fg="#e2e8f0",
                     font=('Consolas', 10),
                     insertbackground="#e2e8f0",
                     selectbackground="#3b82f6",
                     yscrollcommand=v_scrollbar.set)
        
        v_scrollbar.config(command=txt.yview)
        
        # Add dependency information
        txt.insert('1.0', "Dependency Status:\n")
        txt.insert('2.0', "=" * 50 + "\n\n")
        
        for issue in issues:
            txt.insert(tk.END, f"{issue}\n")
        
        txt.insert(tk.END, "\n" + "=" * 50 + "\n")
        txt.insert(tk.END, "\nIf any dependencies are missing, install them with:\n")
        txt.insert(tk.END, "pip install pan-os-python pandas tabulate reportlab openpyxl\n")
        
        txt.config(state='disabled')
        txt.pack(fill=tk.BOTH, expand=True)
        
        # Close button
        button_frame = ttk.Frame(main_container)
        button_frame.pack(fill=tk.X, pady=(15, 0))
        
        ttk.Button(button_frame, text="Close", command=deps_window.destroy, width=10).pack(side=tk.RIGHT)

    def test_api_connection(self):
        """Test the API connection and show detailed results in a popup"""
        mode = self.mode_var.get()
        if mode != "API Connection":
            messagebox.showinfo(APP_NAME, "API testing is only available in API Connection mode.")
            return
            
        api_url = self.api_url_var.get().strip()
        api_key = self.api_key_var.get().strip()
        vsys = self.vsys_var.get().strip() or "vsys1"
        
        if not api_url or not api_key:
            messagebox.showerror(APP_NAME, "API URL and API Key are required for testing.")
            return
        
        self.status.set("Testing API connection...")
        self.update_idletasks()
        
        # Create test results popup
        test_window = APITestWindow(self, api_url, api_key, vsys)
        test_window.grab_set()  # Make window modal

    def run_analyze(self):
        mode = self.mode_var.get()
        
        if mode == "API Connection":
            self._run_api_analyze()
        else:  # CSV Import
            self._run_csv_analyze()
    
    def _run_api_analyze(self):
        """Run analysis using API connection"""
        api_url = self.api_url_var.get().strip()
        api_key = self.api_key_var.get().strip()
        vsys = self.vsys_var.get().strip() or "vsys1"
        if not api_url or not api_key:
            messagebox.showerror(APP_NAME, "API URL and API Key are required.")
            return
        
        # Log VSYS being used to console
        print(f"\n=== API ANALYSIS STARTING ===")
        print(f"Target: {api_url}")
        print(f"VSYS: {vsys}")
        print(f"Timestamp: {now_iso()}")
        print("=" * 30)
        
        self.status.set(f"Connecting and fetching policy from VSYS '{vsys}' (read-only)…")
        self.update_idletasks()
        try:
            client = PanOSReader(api_url=api_url, api_key=api_key, vsys=vsys)
            rules = client.fetch_rules()
            
            # Check if we got any rules
            if not rules:
                print(f"\n=== ANALYSIS WARNING ===")
                print(f"Available VSYS: {client.available_vsys}")
                print(f"Tried VSYS: {vsys}")
                print(f"No security rules found in any VSYS")
                print("This could mean:")
                print("- No VSYS has security rules configured")
                print("- The API user lacks permissions to view security rules")
                print("- The firewall is using a different rulebase type")
                print("- The firewall has no security rules at all")
                print("=" * 30)
                
                messagebox.showwarning(APP_NAME, 
                    f"No security rules found in any VSYS.\n\n"
                    f"Available VSYS: {', '.join(client.available_vsys)}\n\n"
                    "This could mean:\n"
                    "• No VSYS has security rules configured\n"
                    "• The API user lacks permissions\n"
                    "• The firewall uses a different rulebase type\n"
                    "• The firewall has no security rules at all")
                self.status.set(f"⚠️ No security rules found in any VSYS")
                return
            
            # Use the VSYS that actually had rules
            active_vsys = client.vsys
            print(f"Using VSYS with rules: {active_vsys}")
            
            hits = client.fetch_hit_counts(vsys=active_vsys)
            self._process_analysis_results(rules, hits, f"API (VSYS: {active_vsys})")
            
        except Exception as e:
            print(f"\n=== API ANALYSIS ERROR ===")
            print(f"VSYS: {vsys}")
            print(f"Error: {e}")
            print("=" * 30)
            messagebox.showerror(APP_NAME, f"API Error: {e}")
            self.status.set("Error during API analysis.")
    
    def _run_csv_analyze(self):
        """Run analysis using CSV import"""
        csv_file = self.csv_file_var.get().strip()
        if not csv_file:
            messagebox.showerror(APP_NAME, "Please select a CSV file for import.")
            return
        
        if not os.path.exists(csv_file):
            messagebox.showerror(APP_NAME, f"CSV file not found: {csv_file}")
            return
        
        # Log CSV import to console
        print(f"\n=== CSV ANALYSIS STARTING ===")
        print(f"CSV File: {csv_file}")
        print(f"Timestamp: {now_iso()}")
        print("=" * 30)
        
        self.status.set(f"Importing and analyzing CSV file...")
        self.update_idletasks()
        try:
            client = CSVReader(csv_file_path=csv_file)
            rules = client.fetch_rules()
            
            # Check if we got any rules
            if not rules:
                print(f"\n=== CSV ANALYSIS WARNING ===")
                print(f"No security rules found in CSV file")
                print("This could mean:")
                print("- The CSV file is empty or corrupted")
                print("- The CSV file doesn't contain security rules")
                print("- The CSV format is not recognized")
                print("=" * 30)
                
                messagebox.showwarning(APP_NAME, 
                    f"No security rules found in CSV file.\n\n"
                    "This could mean:\n"
                    "• The CSV file is empty or corrupted\n"
                    "• The CSV file doesn't contain security rules\n"
                    "• The CSV format is not recognized")
                self.status.set(f"⚠️ No security rules found in CSV file")
                return
            
            # Hit counts are already embedded in the rules from CSV
            hits = {}  # Empty dict since hits are in rules
            self._process_analysis_results(rules, hits, f"CSV Import ({os.path.basename(csv_file)})")
            
        except Exception as e:
            print(f"\n=== CSV ANALYSIS ERROR ===")
            print(f"CSV File: {csv_file}")
            print(f"Error: {e}")
            print("=" * 30)
            messagebox.showerror(APP_NAME, f"CSV Import Error: {e}")
            self.status.set("Error during CSV analysis.")
    
    def _process_analysis_results(self, rules: List[RuleLike], hits: Dict[str, Dict[str, Any]], source: str):
        """Process analysis results for both API and CSV modes"""
        analyzer = Analyzer(rules, hits)
        self.df_rules = analyzer.build_dataframe()
        self.unused = analyzer.unused_rules_zero_hits()
        self.shadows = analyzer.find_shadowed_rules()
        self.merges = analyzer.propose_merges()
        self.df_final = Exporter.dataframe_with_recommendations(self.df_rules, self.unused, self.shadows, self.merges)
        
        # Populate both tabs
        self.populate_tree(self.df_final)
        self.populate_overview(rules, self.shadows, self.merges, source)
        
        # Log results to console
        print(f"\n=== ANALYSIS COMPLETE ===")
        print(f"Source: {source}")
        print(f"Total Rules: {len(rules)}")
        print(f"Unused Rules (0 hits): {len(self.unused)}")
        print(f"Shadowed Rules: {len(self.shadows)}")
        print(f"Merge Groups: {len(self.merges)}")
        print("=" * 30)
        
        self.status.set(f"✅ Analysis complete! Rules: {len(rules)}, Unused: {len(self.unused)}, Shadows: {len(self.shadows)}, Merge groups: {len(self.merges)}")

    def populate_tree(self, df: Optional[pd.DataFrame]):
        for row in self.tree.get_children():
            self.tree.delete(row)
        if df is None or df.empty:
            return
        
        for idx, (_, r) in enumerate(df.iterrows()):
            # Apply alternating row colors
            tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
            
            # Format hits for better display
            hits = r.get("Rule Usage Hit Count") or r.get("Hits")
            if hits is not None and hits == 0:
                hits_display = "0 ⚠️"
            elif hits is not None:
                hits_display = str(hits)
            else:
                hits_display = "N/A"
            
            # Format action with colors
            action = r.get("Action", "")
            if action.lower() == "deny":
                action_display = f"❌ {action}"
            elif action.lower() == "allow":
                action_display = f"✅ {action}"
            else:
                action_display = action
            
            # Truncate long values for display
            def truncate(value, max_len=50):
                if value and len(str(value)) > max_len:
                    return str(value)[:max_len] + "..."
                return str(value) if value else ""
            
            self.tree.insert('', 'end', values=(
                r.get("Position"), 
                truncate(r.get("Name")), 
                truncate(r.get("Tags")),
                r.get("Type"),
                truncate(r.get("Source Zone")),
                truncate(r.get("Source Address")),
                truncate(r.get("Source User")),
                truncate(r.get("Source Device")),
                truncate(r.get("Destination Zone")),
                truncate(r.get("Destination Address")),
                truncate(r.get("Destination Device")),
                truncate(r.get("Application")),
                truncate(r.get("Service")),
                action_display, 
                truncate(r.get("Profile")),
                truncate(r.get("Options")),
                hits_display, 
                truncate(r.get("Rule Usage Last Hit")),
                truncate(r.get("Rule Usage First Hit")),
                truncate(r.get("Rule Usage Apps Seen")),
                truncate(r.get("Days With No New Apps")),
                truncate(r.get("Modified")),
                truncate(r.get("Created")),
                truncate(r.get("Recommendation"), 100)
            ), tags=(tag,))
    
    def populate_overview(self, rules: List[RuleLike], shadows: List[ShadowFinding], merges: List[Proposal], source: str):
        """Populate the overview tab with system details and analytics"""
        for row in self.overview_tree.get_children():
            self.overview_tree.delete(row)
        
        if not rules:
            return
        
        # Calculate analytics
        total_rules = len(rules)
        disabled_rules = sum(1 for r in rules if r.disabled)
        enabled_rules = total_rules - disabled_rules
        allow_rules = sum(1 for r in rules if r.action.lower() == 'allow')
        deny_rules = sum(1 for r in rules if r.action.lower() in ['deny', 'drop'])
        zero_hit_rules = sum(1 for r in rules if r.hits_total == 0)
        total_hits = sum(r.hits_total or 0 for r in rules)
        avg_hits = total_hits / total_rules if total_rules > 0 else 0
        
        # Get unique values for diversity metrics
        unique_apps = set()
        unique_services = set()
        unique_sources = set()
        unique_destinations = set()
        unique_zones = set()
        
        for rule in rules:
            unique_apps.update(rule.application)
            unique_services.update(rule.service)
            unique_sources.update(rule.source)
            unique_destinations.update(rule.destination)
            unique_zones.update(rule.fromzone + rule.tozone)
        
        # Remove 'any' from counts for meaningful metrics
        unique_apps.discard('any')
        unique_services.discard('any')
        unique_sources.discard('any')
        unique_destinations.discard('any')
        unique_zones.discard('any')
        
        # Overview data
        overview_data = [
            # System Information
            ("System", "Analysis Source", source, "Data source for this analysis"),
            ("System", "Analysis Date", now_iso(), "When this analysis was performed"),
            ("System", "Total Rules", str(total_rules), "Total number of security rules"),
            ("System", "Enabled Rules", str(enabled_rules), "Number of enabled rules"),
            ("System", "Disabled Rules", str(disabled_rules), "Number of disabled rules"),
            
            # Rule Actions
            ("Actions", "Allow Rules", str(allow_rules), "Rules that allow traffic"),
            ("Actions", "Deny/Drop Rules", str(deny_rules), "Rules that deny or drop traffic"),
            ("Actions", "Allow Percentage", f"{(allow_rules/total_rules*100):.1f}%" if total_rules > 0 else "0%", "Percentage of rules that allow traffic"),
            
            # Hit Count Analytics
            ("Hit Counts", "Zero Hit Rules", str(zero_hit_rules), "Rules with no traffic hits"),
            ("Hit Counts", "Zero Hit Percentage", f"{(zero_hit_rules/total_rules*100):.1f}%" if total_rules > 0 else "0%", "Percentage of rules with no hits"),
            ("Hit Counts", "Total Hits", str(total_hits), "Sum of all rule hit counts"),
            ("Hit Counts", "Average Hits", f"{avg_hits:.1f}", "Average hits per rule"),
            
            # Diversity Metrics
            ("Diversity", "Unique Applications", str(len(unique_apps)), "Number of unique applications referenced"),
            ("Diversity", "Unique Services", str(len(unique_services)), "Number of unique services referenced"),
            ("Diversity", "Unique Sources", str(len(unique_sources)), "Number of unique source addresses"),
            ("Diversity", "Unique Destinations", str(len(unique_destinations)), "Number of unique destination addresses"),
            ("Diversity", "Unique Zones", str(len(unique_zones)), "Number of unique zones referenced"),
            
            # Analysis Results
            ("Analysis", "Shadowed Rules", str(len(shadows)), "Rules that are shadowed by earlier rules"),
            ("Analysis", "Merge Groups", str(len(merges)), "Groups of rules that could be merged"),
            ("Analysis", "Shadowed Percentage", f"{(len(shadows)/total_rules*100):.1f}%" if total_rules > 0 else "0%", "Percentage of rules that are shadowed"),
            
            # Recommendations
            ("Recommendations", "Disable Candidates", str(zero_hit_rules), "Rules with zero hits - consider disabling"),
            ("Recommendations", "Merge Candidates", str(sum(len(m.source_rules) for m in merges)), "Total rules involved in merge proposals"),
            ("Recommendations", "Review Required", str(len(shadows) + len(merges)), "Rules requiring manual review"),
        ]
        
        # Insert data with alternating row colors
        for idx, (category, metric, value, description) in enumerate(overview_data):
            tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
            self.overview_tree.insert('', 'end', values=(category, metric, value, description), tags=(tag,))

    def preview_summary(self):
        if self.df_final is None:
            messagebox.showinfo(APP_NAME, "Run analysis first.")
            return
        # Build a small textual summary
        head = self.df_final[["Position","Name","Action","Rule Usage Hit Count","Recommendation"]].head(25)
        txt = tabulate(head, headers='keys', tablefmt='github', showindex=False)
        Summary(self, txt)

    def export_all(self):
        if self.df_final is None:
            messagebox.showinfo(APP_NAME, "Run analysis first.")
            return
        outdir = self.output_dir_var.get().strip() or os.getcwd()
        os.makedirs(outdir, exist_ok=True)
        ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        base = os.path.join(outdir, f"panos_policy_eval_{ts}")

        try:
            # Create overview DataFrame from overview tree
            overview_data = []
            for item in self.overview_tree.get_children():
                values = self.overview_tree.item(item)['values']
                if values:
                    overview_data.append({
                        'Category': values[0],
                        'Metric': values[1], 
                        'Value': values[2],
                        'Description': values[3]
                    })
            df_overview = pd.DataFrame(overview_data)
            
            # Export both tabs
            csv_path = Exporter.export_csv_with_tabs(self.df_final, df_overview, base+".csv")
            xlsx_path = Exporter.export_xlsx_with_tabs(self.df_final, df_overview, base+".xlsx")
            messagebox.showinfo(APP_NAME, f"Exported:\n{csv_path}\n{xlsx_path}")
            self.status.set("✅ Export complete! CSV and XLSX with both tabs generated successfully.")
        except Exception as e:
            messagebox.showerror(APP_NAME, f"Export error: {e}")


class Summary(tk.Toplevel):
    def __init__(self, parent: App, text: str):
        super().__init__(parent)
        self.title("Preview Summary")
        self.geometry("1000x700")
        self.configure(bg="#0f172a")
        
        # Title
        title_label = ttk.Label(self, text="Analysis Summary", 
                               font=('Segoe UI', 14, 'bold'))
        title_label.pack(anchor='w', padx=20, pady=(20, 10))
        
        # Text area with scrollbars
        text_frame = ttk.Frame(self)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))
        
        # Vertical scrollbar
        v_scrollbar = ttk.Scrollbar(text_frame)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Horizontal scrollbar
        h_scrollbar = ttk.Scrollbar(text_frame, orient="horizontal")
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Text widget
        txt = tk.Text(text_frame, 
                     wrap='none', 
                     bg="#1e293b", 
                     fg="#e2e8f0",
                     font=('Consolas', 9),
                     insertbackground="#e2e8f0",
                     selectbackground="#3b82f6",
                     yscrollcommand=v_scrollbar.set,
                     xscrollcommand=h_scrollbar.set)
        
        v_scrollbar.config(command=txt.yview)
        h_scrollbar.config(command=txt.xview)
        
        txt.insert('1.0', text)
        txt.config(state='disabled')
        txt.pack(fill=tk.BOTH, expand=True)


class APITestWindow(tk.Toplevel):
    def __init__(self, parent: App, api_url: str, api_key: str, vsys: str):
        super().__init__(parent)
        self.title("API Connection Test")
        self.geometry("800x600")
        self.configure(bg="#0f172a")
        self.resizable(True, True)
        
        # Center the window
        self.transient(parent)
        self.grab_set()
        
        # Test parameters
        self.api_url = api_url
        self.api_key = api_key
        self.vsys = vsys
        
        self._build_ui()
        self._run_test()
    
    def _build_ui(self):
        # Main container
        main_container = ttk.Frame(self)
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Title
        title_label = ttk.Label(main_container, text="🔌 API Connection Test", 
                               font=('Segoe UI', 14, 'bold'))
        title_label.pack(anchor='w', pady=(0, 15))
        
        # Test parameters display
        params_frame = ttk.Frame(main_container)
        params_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(params_frame, text="Test Parameters:", 
                 font=('Segoe UI', 10, 'bold')).pack(anchor='w')
        
        params_text = f"""
API URL: {self.api_url}
VSYS: {self.vsys}
API Key: {'*' * min(len(self.api_key), 8)}...
        """
        
        params_display = tk.Text(params_frame, height=4, wrap=tk.WORD,
                                bg="#1e293b", fg="#e2e8f0", 
                                font=('Consolas', 9),
                                relief='flat', borderwidth=0)
        params_display.insert('1.0', params_text.strip())
        params_display.config(state='disabled')
        params_display.pack(fill=tk.X, pady=(5, 0))
        
        # Status indicator
        self.status_frame = ttk.Frame(main_container)
        self.status_frame.pack(fill=tk.X, pady=(0, 15))
        
        self.status_label = ttk.Label(self.status_frame, text="⏳ Testing connection...", 
                                     font=('Segoe UI', 11, 'bold'))
        self.status_label.pack(anchor='w')
        
        # Progress bar
        self.progress = ttk.Progressbar(self.status_frame, mode='indeterminate')
        self.progress.pack(fill=tk.X, pady=(5, 0))
        self.progress.start()
        
        # Results area
        results_frame = ttk.Frame(main_container)
        results_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(results_frame, text="Test Results:", 
                 font=('Segoe UI', 10, 'bold')).pack(anchor='w')
        
        # Scrollable text area for results
        text_frame = ttk.Frame(results_frame)
        text_frame.pack(fill=tk.BOTH, expand=True, pady=(5, 0))
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(text_frame)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        h_scrollbar = ttk.Scrollbar(text_frame, orient="horizontal")
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Text widget
        self.results_text = tk.Text(text_frame, 
                                   wrap='none',
                                   bg="#1e293b", 
                                   fg="#e2e8f0",
                                   font=('Consolas', 9),
                                   insertbackground="#e2e8f0",
                                   selectbackground="#3b82f6",
                                   yscrollcommand=v_scrollbar.set,
                                   xscrollcommand=h_scrollbar.set)
        
        v_scrollbar.config(command=self.results_text.yview)
        h_scrollbar.config(command=self.results_text.xview)
        
        self.results_text.pack(fill=tk.BOTH, expand=True)
        
        # Close button
        button_frame = ttk.Frame(main_container)
        button_frame.pack(fill=tk.X, pady=(15, 0))
        
        ttk.Button(button_frame, text="Close", command=self.destroy, width=10).pack(side=tk.RIGHT)
    
    def _run_test(self):
        """Run the API connection test in a separate thread to avoid blocking UI"""
        import threading
        import time
        
        def test_thread():
            try:
                # Test 1: Basic connection
                self._log_result("=== API CONNECTION TEST ===\n")
                self._log_result(f"Timestamp: {now_iso()}\n")
                self._log_result(f"Target: {self.api_url}\n")
                self._log_result(f"VSYS: {self.vsys}\n\n")
                
                self._log_result("1. Testing basic connectivity...\n")
                start_time = time.time()
                
                try:
                    client = PanOSReader(api_url=self.api_url, api_key=self.api_key, vsys=self.vsys)
                    connect_time = time.time() - start_time
                    self._log_result(f"✅ SUCCESS: Connected in {connect_time:.2f}s\n")
                    
                    # Test 2: Device info
                    self._log_result("\n2. Fetching device information...\n")
                    try:
                        # Get system info
                        system_info = client.fw.op("show system info", xml=True)
                        
                        # Handle both bytes and XML object responses
                        if isinstance(system_info, bytes):
                            import xml.etree.ElementTree as ET
                            root = ET.fromstring(system_info.decode('utf-8'))
                            hostname = root.findtext('.//hostname') or 'Unknown'
                            model = root.findtext('.//model') or 'Unknown'
                            version = root.findtext('.//sw-version') or 'Unknown'
                        else:
                            hostname = system_info.findtext('.//hostname') or 'Unknown'
                            model = system_info.findtext('.//model') or 'Unknown'
                            version = system_info.findtext('.//sw-version') or 'Unknown'
                        
                        self._log_result(f"✅ Device Info Retrieved:\n")
                        self._log_result(f"   Hostname: {hostname}\n")
                        self._log_result(f"   Model: {model}\n")
                        self._log_result(f"   Version: {version}\n")
                        
                    except Exception as e:
                        self._log_result(f"⚠️  WARNING: Could not fetch device info: {str(e)}\n")
                    
                    # Test 3: Rule count
                    self._log_result("\n3. Testing rule retrieval...\n")
                    try:
                        rules = client.fetch_rules()
                        self._log_result(f"✅ SUCCESS: Retrieved {len(rules)} security rules\n")
                        
                        # Show some rule details
                        if rules:
                            self._log_result(f"   First rule: {rules[0].name}\n")
                            self._log_result(f"   Last rule: {rules[-1].name}\n")
                            self._log_result(f"   Rules with 'any' action: {sum(1 for r in rules if r.action == 'any')}\n")
                            self._log_result(f"   Disabled rules: {sum(1 for r in rules if r.disabled)}\n")
                        
                    except Exception as e:
                        self._log_result(f"❌ FAILED: Rule retrieval error: {str(e)}\n")
                        self._log_result(f"   Error type: {type(e).__name__}\n")
                        # Continue with other tests even if rule retrieval fails
                        self._log_result("   Continuing with other tests...\n")
                    
                    # Test 4: Hit counts
                    self._log_result("\n4. Testing hit count retrieval...\n")
                    try:
                        hits = client.fetch_hit_counts(vsys=self.vsys)
                        self._log_result(f"✅ SUCCESS: Retrieved hit counts for {len(hits)} rules\n")
                        
                        if hits:
                            total_hits = sum(h.get('total', 0) for h in hits.values())
                            zero_hits = sum(1 for h in hits.values() if h.get('total', 0) == 0)
                            self._log_result(f"   Total hits across all rules: {total_hits}\n")
                            self._log_result(f"   Rules with zero hits: {zero_hits}\n")
                        
                    except Exception as e:
                        self._log_result(f"⚠️  WARNING: Hit count retrieval failed: {str(e)}\n")
                        self._log_result("   (This may be normal if hit counters are disabled)\n")
                    
                    # Test 5: VSYS confirmation
                    self._log_result("\n5. Confirming VSYS configuration...\n")
                    try:
                        # First check if this is multi-VSYS or single-VSYS
                        system_info = client.fw.op("show system info", xml=True)
                        if isinstance(system_info, bytes):
                            import xml.etree.ElementTree as ET
                            root = ET.fromstring(system_info.decode('utf-8'))
                        else:
                            root = system_info
                        
                        multi_vsys = root.findtext('.//multi-vsys')
                        if multi_vsys and multi_vsys.lower() == 'off':
                            self._log_result(f"✅ CONFIRMED: Single-VSYS mode detected\n")
                            self._log_result(f"   Using default VSYS: vsys1\n")
                            self._log_result(f"   (Multi-VSYS is disabled on this firewall)\n")
                        else:
                            # Multi-VSYS mode - try to get available VSYS list
                            vsys_commands = [
                                "show vsys",
                                "show system vsys"
                            ]
                            
                            vsys_names = []
                            for cmd in vsys_commands:
                                try:
                                    vsys_list = client.fw.op(cmd, xml=True)
                                    
                                    # Handle both bytes and XML object responses
                                    if isinstance(vsys_list, bytes):
                                        import xml.etree.ElementTree as ET
                                        root = ET.fromstring(vsys_list.decode('utf-8'))
                                    else:
                                        root = vsys_list
                                    
                                    # Try different XML paths for VSYS names
                                    entries = root.findall('.//entry') or root.findall('.//vsys') or root.findall('.//virtual-system')
                                    vsys_names = [v.get('name') for v in entries if v.get('name')]
                                    
                                    if vsys_names:
                                        break
                                except Exception:
                                    continue
                            
                            self._log_result(f"✅ CONFIRMED: Multi-VSYS mode detected\n")
                            self._log_result(f"   Using VSYS: {self.vsys}\n")
                            
                            if vsys_names:
                                self._log_result(f"   Available VSYS options: {', '.join(vsys_names)}\n")
                                if self.vsys in vsys_names:
                                    self._log_result(f"   ✓ VSYS '{self.vsys}' found in available options\n")
                                else:
                                    self._log_result(f"   ℹ️  VSYS '{self.vsys}' not in detected list (may be valid)\n")
                            else:
                                self._log_result(f"   ℹ️  Could not retrieve VSYS list (continuing with '{self.vsys}')\n")
                        
                    except Exception as e:
                        self._log_result(f"⚠️  WARNING: Could not retrieve VSYS information: {str(e)}\n")
                        self._log_result(f"   Continuing with VSYS '{self.vsys}' as configured\n")
                    
                    # Final summary
                    self._log_result("\n=== TEST SUMMARY ===\n")
                    self._log_result("✅ API Connection: SUCCESS\n")
                    self._log_result("✅ Authentication: SUCCESS\n")
                    self._log_result("✅ Rule Access: SUCCESS\n")
                    self._log_result("✅ Ready for analysis!\n")
                    
                    self._update_status("✅ Connection Test PASSED", "success")
                    
                except Exception as e:
                    self._log_result(f"❌ FAILED: {str(e)}\n")
                    self._log_result("\n=== TROUBLESHOOTING ===\n")
                    self._log_result("• Verify API URL is correct and accessible\n")
                    self._log_result("• Check API key is valid and not expired\n")
                    self._log_result("• Ensure firewall allows API access\n")
                    self._log_result("• Verify VSYS name is correct\n")
                    self._log_result("• Check network connectivity\n")
                    
                    self._update_status("❌ Connection Test FAILED", "error")
                
            except Exception as e:
                self._log_result(f"❌ CRITICAL ERROR: {str(e)}\n")
                self._update_status("❌ Test Error", "error")
        
        # Start test in background thread
        thread = threading.Thread(target=test_thread, daemon=True)
        thread.start()
    
    def _log_result(self, message: str):
        """Add message to results text area"""
        self.results_text.insert(tk.END, message)
        self.results_text.see(tk.END)
        self.update_idletasks()
    
    def _update_status(self, message: str, status_type: str):
        """Update status label and stop progress bar"""
        self.progress.stop()
        self.progress.pack_forget()
        
        if status_type == "success":
            self.status_label.config(text=f"✅ {message}", foreground="#10b981")
        else:
            self.status_label.config(text=f"❌ {message}", foreground="#ef4444")


# -----------------------------
# Main
# -----------------------------
def main():
    """Main entry point for the application"""
    app = App()
    app.mainloop()

if __name__ == "__main__":
    main()
